import hashlib
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

from django.db import transaction

from dashboards.models import (
    Company,
    FinancialPeriod,
    FinancialValue,
    ImportChange,
    MetricDefinition,
    RegulatoryFeeValue,
    WorkbookImport,
)

PARSER_VERSION = "mno-bwa-v2"


@dataclass
class ImportResult:
    row_count: int
    change_count: int
    warning_count: int = 0


def normalize_label(value):
    return " ".join(str(value or "").upper().replace("-", " ").split())


def parse_decimal(value, unit_type):
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        parsed = value
    else:
        text = str(value).strip().replace(",", "").replace("GHS", "").replace("%", "")
        if text in {"", "-", "N/A"}:
            return None
        try:
            parsed = Decimal(text)
        except InvalidOperation:
            return None

    if unit_type == MetricDefinition.UnitType.PERCENT and abs(parsed) > 1:
        parsed = parsed / Decimal("100")
    return parsed


def workbook_sha256(file_field):
    hasher = hashlib.sha256()
    with file_field.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def import_workbook(import_record):
    category_slug = import_record.category.slug

    import_record.file_sha256 = workbook_sha256(import_record.workbook_file)
    import_record.parser_version = PARSER_VERSION
    import_record.save(update_fields=["file_sha256", "parser_version", "updated_at"])
    import_record.mark_processing()

    try:
        if category_slug == "bwa":
            result = process_bwa_import(import_record)
        elif category_slug == "mobile-network":
            result = process_mobile_network_import(import_record)
        else:
            raise NotImplementedError(
                f"Workbook imports for category '{category_slug}' are not supported."
            )
    except Exception as exc:
        import_record.mark_failed(exc)
        raise

    import_record.mark_completed(row_count=result.row_count, warning_count=result.warning_count)
    return result


def process_bwa_import(import_record):
    from openpyxl import load_workbook

    with import_record.workbook_file.open("rb") as handle:
        workbook = load_workbook(handle, data_only=True, read_only=True)
        # Check if sheets are year-named
        year_sheets = [s for s in workbook.sheetnames if s.isdigit() and len(s) == 4]
        if not year_sheets:
            # Check if it's the BWA quarterly report sheet
            if any("BWA Subs" in s for s in workbook.sheetnames):
                return process_bwa_quarterly_report(import_record, workbook)
            raise ValueError(
                "No sheets named with a 4-digit year or quarterly operator sheets found."
            )

        rows = parse_bwa_workbook(import_record, workbook)
        return persist_financial_rows(import_record, rows)


def parse_bwa_workbook(import_record, workbook):
    metrics = list(MetricDefinition.objects.filter(category=import_record.category, is_active=True))
    if not metrics:
        raise ValueError("No metric definitions exist for this category. Run seed_reference_data.")

    parsed_rows = []
    for sheet_name in workbook.sheetnames:
        if not str(sheet_name).isdigit() or len(str(sheet_name)) != 4:
            continue
        year = int(sheet_name)
        sheet = workbook[sheet_name]
        sheet_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        for metric in metrics:
            value = find_metric_value(sheet_rows, metric)
            if value is not None:
                parsed_rows.append(
                    {
                        "year": year,
                        "company_slug": "telesol",
                        "metric": metric,
                        "value": value,
                    }
                )
    if not parsed_rows:
        raise ValueError("No values were parsed. Check that metric labels match.")
    return parsed_rows


def process_bwa_quarterly_report(import_record, workbook):
    # This is a quarterly subscriber report.
    # We will parse subscriptions from "20. BWA Subs per Operator"
    # and map the latest quarter (e.g. Q3/Q4) or general subscription value to the corresponding year.
    sheet_name = "20. BWA Subs per Operator"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Required sheet '{sheet_name}' not found in quarterly BWA workbook.")

    sheet = workbook[sheet_name]
    sheet_rows = [list(row) for row in sheet.iter_rows(values_only=True)]

    # Look for headers row to find quarters (e.g. Q4 2023, Q1 2024, etc.)
    headers = []
    header_idx = -1
    for i, row in enumerate(sheet_rows):
        cleaned = [str(c).strip().lower() for c in row if c is not None]
        if "bwa operator" in cleaned:
            headers = [str(c).strip() for c in row]
            header_idx = i
            break

    if header_idx == -1:
        raise ValueError("Operator subscription headers not found.")

    companies = {company.name.lower(): company for company in Company.objects.all()}
    metrics = {metric.key: metric for metric in MetricDefinition.objects.filter(category=import_record.category)}

    # We need a subscriber count metric definition
    # If not found, we will map it to 'count' or create a dummy metric definition
    metric, _ = MetricDefinition.objects.get_or_create(
        category=import_record.category,
        key="subscribers",
        defaults={
            "label": "Subscribers",
            "unit_type": MetricDefinition.UnitType.COUNT,
        }
    )

    parsed_rows = []
    # Scan operator rows below headers
    for row in sheet_rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        op_name = str(row[0]).strip().lower()
        if op_name in {"industry total", "total"}:
            continue

        company = companies.get(op_name)
        if not company:
            # Auto-create company if missing
            company = Company.objects.create(slug=op_name.replace(" ", "-"), name=op_name.title())
            companies[op_name] = company

        # Parse each quarter column
        for col_idx in range(2, len(row)):
            col_header = headers[col_idx]
            # e.g. "Q3 2024" or "2024"
            parts = col_header.split()
            year = None
            for p in parts:
                if p.isdigit() and len(p) == 4:
                    year = int(p)
                    break
            if year:
                val = parse_decimal(row[col_idx], MetricDefinition.UnitType.COUNT)
                if val is not None:
                    parsed_rows.append({
                        "year": year,
                        "company_slug": company.slug,
                        "metric": metric,
                        "value": val
                    })

    return persist_financial_rows(import_record, parsed_rows)


def process_mobile_network_import(import_record):
    from openpyxl import load_workbook

    with import_record.workbook_file.open("rb") as handle:
        workbook = load_workbook(handle, data_only=True, read_only=True)
        # Check if it is the 1% Regulatory Fee sheet
        is_regulatory_fee = "SUMMARY" in workbook.sheetnames or any(
            "regulatory fee" in str(s).lower() for s in workbook.sheetnames
        )

        if is_regulatory_fee:
            return process_regulatory_fees(import_record, workbook)
        else:
            return process_mno_financials(import_record, workbook)


def process_mno_financials(import_record, workbook):
    metrics = list(MetricDefinition.objects.filter(category=import_record.category, is_active=True))
    if not metrics:
        raise ValueError("No metric definitions exist for this category. Run seed_reference_data.")

    companies = {company.slug: company for company in Company.objects.filter(is_active=True)}

    parsed_rows = []
    for sheet_name in workbook.sheetnames:
        if not str(sheet_name).isdigit() or len(str(sheet_name)) != 4:
            continue
        year = int(sheet_name)
        sheet = workbook[sheet_name]
        sheet_rows = [list(row) for row in sheet.iter_rows(values_only=True)]

        # Find operator headers in first 5 rows
        headers = []
        for r in sheet_rows[:5]:
            cleaned = [str(cell).strip().lower() if cell is not None else "" for cell in r]
            if "mtn" in cleaned or "telecel" in cleaned:
                headers = cleaned
                break

        if not headers:
            continue

        # Map operator column indices
        company_cols = {}
        for idx, col_name in enumerate(headers):
            if col_name in companies and col_name not in company_cols:
                company_cols[col_name] = idx

        for metric in metrics:
            metric_row = find_metric_row_in_sheet(sheet_rows, metric)
            if metric_row is not None:
                for comp_slug, col_idx in company_cols.items():
                    if col_idx < len(metric_row):
                        val = parse_decimal(metric_row[col_idx], metric.unit_type)
                        if val is not None:
                            parsed_rows.append(
                                {
                                    "year": year,
                                    "company_slug": comp_slug,
                                    "metric": metric,
                                    "value": val,
                                }
                            )

    if not parsed_rows:
        raise ValueError("No operator financial values were parsed.")

    return persist_financial_rows(import_record, parsed_rows)


def process_regulatory_fees(import_record, workbook):
    companies = {company.name.lower(): company for company in Company.objects.filter(is_active=True)}
    # Add mapping for AT (which is styled as "AT" but can match "at" or "at (airteltigo)")
    companies["at (airteltigo)"] = companies.get("at")
    companies["airteltigo"] = companies.get("at")

    parsed_fees = []
    for sheet_name in workbook.sheetnames:
        if not str(sheet_name).isdigit() or len(str(sheet_name)) != 4:
            continue
        year = int(sheet_name)
        sheet = workbook[sheet_name]
        sheet_rows = [list(row) for row in sheet.iter_rows(values_only=True)]

        current_company = None
        for row in sheet_rows:
            cleaned = [str(c).strip().lower() if c is not None else "" for c in row]
            # Column index 3 is OPERATOR
            if len(cleaned) > 3:
                op_val = cleaned[3]
                if op_val in companies:
                    current_company = companies[op_val]

            # Column index 2 is PERIOD (Q1, Q2, Q3, Q4)
            if len(cleaned) > 2 and current_company:
                period_val = cleaned[2].upper()
                if period_val in {"Q1", "Q2", "Q3", "Q4"}:
                    invoice_val = parse_decimal(row[4], MetricDefinition.UnitType.MONEY) or Decimal("0")
                    payment_val = parse_decimal(row[5], MetricDefinition.UnitType.MONEY) or Decimal("0")
                    parsed_fees.append(
                        {
                            "year": year,
                            "company": current_company,
                            "invoice_issued": invoice_val,
                            "payment_received": payment_val,
                        }
                    )

    if not parsed_fees:
        raise ValueError("No regulatory fee values were parsed.")

    # Aggregate by company and year
    aggregated = {}
    for item in parsed_fees:
        key = (item["year"], item["company"])
        if key not in aggregated:
            aggregated[key] = {
                "invoice_issued": Decimal("0"),
                "payment_received": Decimal("0"),
            }
        aggregated[key]["invoice_issued"] += item["invoice_issued"]
        aggregated[key]["payment_received"] += item["payment_received"]

    change_count = 0
    row_count = 0

    with transaction.atomic():
        for (year, company), vals in aggregated.items():
            period, _ = FinancialPeriod.objects.get_or_create(year=year)
            outstanding = vals["invoice_issued"] - vals["payment_received"]

            existing = RegulatoryFeeValue.objects.filter(
                category=import_record.category,
                company=company,
                period=period,
            ).first()

            val_changed = (
                existing is None
                or existing.invoice_issued != vals["invoice_issued"]
                or existing.payment_received != vals["payment_received"]
            )

            RegulatoryFeeValue.objects.update_or_create(
                category=import_record.category,
                company=company,
                period=period,
                defaults={
                    "invoice_issued": vals["invoice_issued"],
                    "payment_received": vals["payment_received"],
                    "outstanding": outstanding,
                    "fee_to_revenue": Decimal("0"),
                    "source_import": import_record,
                },
            )

            if val_changed:
                change_count += 1
            row_count += 1

    return ImportResult(row_count=row_count, change_count=change_count)


def find_metric_value(sheet_rows, metric):
    aliases = [normalize_label(metric.label)] + [normalize_label(alias) for alias in metric.aliases]
    for row in sheet_rows:
        normalized_cells = [normalize_label(cell) for cell in row]
        for index, label in enumerate(normalized_cells):
            if label in aliases:
                for candidate in row[index + 1 :]:
                    value = parse_decimal(candidate, metric.unit_type)
                    if value is not None:
                        return value
    return None


def find_metric_row_in_sheet(sheet_rows, metric):
    aliases = [normalize_label(metric.label)] + [normalize_label(alias) for alias in metric.aliases]
    for row in sheet_rows:
        normalized_cells = [normalize_label(cell) for cell in row]
        for index, label in enumerate(normalized_cells):
            if label in aliases:
                return row
    return None


@transaction.atomic
def persist_financial_rows(import_record, rows):
    company_cache = {company.slug: company for company in Company.objects.all()}
    change_count = 0

    for row in rows:
        period, _ = FinancialPeriod.objects.get_or_create(year=row["year"])
        company = company_cache.get(row["company_slug"])
        if company is None:
            raise ValueError(f"Unknown company slug: {row['company_slug']}")

        existing = FinancialValue.objects.filter(
            category=import_record.category,
            company=company,
            period=period,
            metric=row["metric"],
        ).first()

        old_value = existing.value if existing else None
        value_changed = old_value != row["value"]

        financial_value, _ = FinancialValue.objects.update_or_create(
            category=import_record.category,
            company=company,
            period=period,
            metric=row["metric"],
            defaults={"value": row["value"], "source_import": import_record},
        )

        if value_changed:
            ImportChange.objects.create(
                import_record=import_record,
                category=import_record.category,
                company=company,
                period=period,
                metric=row["metric"],
                old_value=old_value,
                new_value=financial_value.value,
            )
            change_count += 1

    return ImportResult(row_count=len(rows), change_count=change_count)
